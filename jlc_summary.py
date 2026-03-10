import os
import sys
import json
import glob
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 全局变量用于存储汇总数据
all_accounts = []
summary_info = {
    'total_groups': 0,
    'total_accounts': 0,
    'success_accounts': 0,
    'failed_accounts': [],
    'password_error_accounts': []
}

def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)

def parse_result_file(filepath):
    """解析单个账号组的结果文件"""
    accounts = []
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
            data = json.loads(content)
            
            group_index = data.get('group_index', 0)
            group_accounts = data.get('accounts', [])
            
            for acc in group_accounts:
                acc['group_index'] = group_index
                accounts.append(acc)
                
                # 更新汇总信息
                if acc.get('password_error', False):
                    summary_info['password_error_accounts'].append({
                        'index': acc.get('account_index', 0),
                        'group': group_index,
                        'username': acc.get('username', '')
                    })
                elif not acc.get('jindou_success', False):
                    summary_info['failed_accounts'].append({
                        'index': acc.get('account_index', 0),
                        'group': group_index,
                        'username': acc.get('username', '')
                    })
                else:
                    summary_info['success_accounts'] += 1
                    
            summary_info['total_accounts'] += len(group_accounts)
            summary_info['total_groups'] = max(summary_info['total_groups'], group_index)
            
    except Exception as e:
        log(f"解析文件 {filepath} 失败: {e}")
    
    return accounts

def collect_all_results():
    """收集所有账号组的结果文件"""
    global all_accounts
    
    # 查找所有结果文件
    result_files = glob.glob('jlc_result_*.json')
    
    if not result_files:
        log("未找到任何结果文件")
        return False
    
    log(f"找到 {len(result_files)} 个结果文件")
    
    for filepath in result_files:
        accounts = parse_result_file(filepath)
        all_accounts.extend(accounts)
        log(f"从 {filepath} 解析了 {len(accounts)} 个账号")
    
    # 计算每个账号的显示金豆（取 initial_jindou 和 final_jindou 的最大值，防止失败时为0）
    for acc in all_accounts:
        acc['display_jindou'] = max(acc.get('final_jindou', 0), acc.get('initial_jindou', 0))
    
    # 按显示金豆数量排序（由高到低）
    all_accounts.sort(key=lambda x: x.get('display_jindou', 0), reverse=True)
    
    return True

def calculate_year_end_prediction(current_beans):
    """计算年底金豆预测数量"""
    try:
        now = datetime.now()
        year_end = datetime(now.year, 12, 31)
        remaining_days = (year_end - now).days
        if remaining_days < 0:
            remaining_days = 0
        # 按照一周大约22个金豆计算，每天平均约 22/7 个
        estimated_future_beans = int(remaining_days * (22 / 7))
        return int(current_beans) + estimated_future_beans
    except Exception:
        return int(current_beans)

def get_display_status(acc):
    """获取签到状态的显示文本"""
    if acc.get('password_error', False):
        return '密码错误'
    
    jindou_status = acc.get('jindou_status', '未知')
    jindou_success = acc.get('jindou_success', False)
    
    if jindou_success:
        # 成功的状态统一显示
        if jindou_status == '已签到过':
            return '已签到过'
        elif jindou_status == '领取奖励成功':
            return '签到成功(有奖励)'
        elif jindou_status == '签到成功':
            return '签到成功'
        else:
            return '签到成功'
    else:
        # 失败的状态直接显示原因
        return jindou_status

def generate_excel():
    """生成Excel排名文件"""
    # 获取当前日期
    now = datetime.now()
    month = now.month
    day = now.day
    
    filename = f"{month}.{day}立创金豆排名.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "金豆排名"
    
    # 设置标题行
    headers = ['排名', '金豆数量', '客编', '密码', '归属账号组', '签到状态', '年底预计']
    ws.append(headers)
    
    # 设置标题行样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 定义签到状态的颜色样式
    status_styles = {
        'success': Font(color="32CD32", bold=True),      # 绿色 - 成功
        'already': Font(color="32CD32", bold=True),       # 绿色 - 已签到过
        'fail': Font(color="C00000", bold=True),          # 深红色 - 失败
        'password': Font(color="FF6600", bold=True),      # 橙色 - 密码错误
    }
    
    total_cols = len(headers)  # 7列
    
    # 先记录每个账号的全局排名，供后续 Sheet 使用
    for rank, acc in enumerate(all_accounts, 1):
        acc['global_rank'] = rank
    
    # ==============================
    # 填充第一个 Sheet（全局金豆排名）
    # ==============================
    for rank, acc in enumerate(all_accounts, 1):
        username = acc.get('username', '')
        display_jindou = acc.get('display_jindou', max(acc.get('final_jindou', 0), acc.get('initial_jindou', 0)))
        actual_password = acc.get('actual_password', '')
        group_index = acc.get('group_index', 0)
        jindou_success = acc.get('jindou_success', False)
        password_error = acc.get('password_error', False)
        
        display_password = actual_password if actual_password else ''
        display_status = get_display_status(acc)
        
        # 年底预计（强制转为int）
        if display_jindou > 0:
            year_end_prediction = calculate_year_end_prediction(int(display_jindou))
        else:
            year_end_prediction = ''
        
        row_data = [
            rank,
            display_jindou,
            username,
            display_password,
            f"{group_index}组账号{acc.get('account_index', 0)}",
            display_status,
            year_end_prediction
        ]
        ws.append(row_data)
        
        # 设置数据行样式
        row_num = rank + 1
        for col_num in range(1, total_cols + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 排名列加粗
            if col_num == 1:
                cell.font = Font(bold=True)
            
            # 金豆数量列使用不同颜色区分
            if col_num == 2:
                if display_jindou >= 500:
                    cell.font = Font(color="C00000", bold=True)  # 深红色
                elif display_jindou >= 300:
                    cell.font = Font(color="FF6600", bold=True)  # 橙色
                elif display_jindou >= 100:
                    cell.font = Font(color="0070C0", bold=True)  # 蓝色
            
            # 签到状态列样式
            if col_num == 6:
                if password_error:
                    cell.font = status_styles['password']
                elif jindou_success:
                    if display_status == '已签到过':
                        cell.font = status_styles['already']
                    else:
                        cell.font = status_styles['success']
                else:
                    cell.font = status_styles['fail']
            
            # 年底预计列样式
            if col_num == 7 and isinstance(year_end_prediction, (int, float)):
                if year_end_prediction >= 1500:
                    cell.font = Font(color="C00000", bold=True)
                elif year_end_prediction >= 1000:
                    cell.font = Font(color="FF6600", bold=True)
                elif year_end_prediction >= 800:
                    cell.font = Font(color="0070C0", bold=True)
    
    # ==============================
    # 填充第二个 Sheet（按组及原顺序排序）
    # ==============================
    ws_group = wb.create_sheet("按组顺序")
    ws_group.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_group.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 按照组别和原本的账号顺序排序
    accounts_by_group = sorted(all_accounts, key=lambda x: (x.get('group_index', 0), x.get('account_index', 0)))
    
    current_row = 2
    last_group = None
    
    for acc in accounts_by_group:
        group_index = acc.get('group_index', 0)
        
        # 分隔行
        if last_group is not None and group_index != last_group:
            ws_group.append([''] * total_cols)
            for col_num in range(1, total_cols + 1):
                cell = ws_group.cell(row=current_row, column=col_num)
                cell.fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
            current_row += 1
            
        last_group = group_index
        
        username = acc.get('username', '')
        display_jindou = acc.get('display_jindou', max(acc.get('final_jindou', 0), acc.get('initial_jindou', 0)))
        actual_password = acc.get('actual_password', '')
        jindou_success = acc.get('jindou_success', False)
        password_error = acc.get('password_error', False)
        
        display_password = actual_password if actual_password else ''
        display_status = get_display_status(acc)
        
        if display_jindou > 0:
            year_end_prediction = calculate_year_end_prediction(int(display_jindou))
        else:
            year_end_prediction = ''
            
        global_rank = acc.get('global_rank', '')
        
        row_data = [
            global_rank, display_jindou, username, display_password,
            f"{group_index}组账号{acc.get('account_index', 0)}", display_status, year_end_prediction
        ]
        ws_group.append(row_data)
        
        for col_num in range(1, total_cols + 1):
            cell = ws_group.cell(row=current_row, column=col_num)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if col_num == 1:
                cell.font = Font(bold=True)
            if col_num == 2:
                if display_jindou >= 500:
                    cell.font = Font(color="C00000", bold=True)
                elif display_jindou >= 300:
                    cell.font = Font(color="FF6600", bold=True)
                elif display_jindou >= 100:
                    cell.font = Font(color="0070C0", bold=True)
            if col_num == 6:
                if password_error:
                    cell.font = status_styles['password']
                elif jindou_success:
                    if display_status == '已签到过':
                        cell.font = status_styles['already']
                    else:
                        cell.font = status_styles['success']
                else:
                    cell.font = status_styles['fail']
            if col_num == 7 and isinstance(year_end_prediction, (int, float)):
                if year_end_prediction >= 1500:
                    cell.font = Font(color="C00000", bold=True)
                elif year_end_prediction >= 1000:
                    cell.font = Font(color="FF6600", bold=True)
                elif year_end_prediction >= 800:
                    cell.font = Font(color="0070C0", bold=True)
                    
        current_row += 1

    # ==============================
    # 填充第三个 Sheet（按组排名 - 组内金豆从高到低）
    # ==============================
    ws_group_rank = wb.create_sheet("按组排名")
    ws_group_rank.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_group_rank.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 按照组别升序，再按金豆数量降序（加负号实现降序）
    accounts_by_group_rank = sorted(
        all_accounts, 
        key=lambda x: (
            x.get('group_index', 0), 
            -x.get('display_jindou', max(x.get('final_jindou', 0), x.get('initial_jindou', 0)))
        )
    )
    
    current_row_rank = 2
    last_group_rank = None
    
    for acc in accounts_by_group_rank:
        group_index = acc.get('group_index', 0)
        
        # 分隔行
        if last_group_rank is not None and group_index != last_group_rank:
            ws_group_rank.append([''] * total_cols)
            for col_num in range(1, total_cols + 1):
                cell = ws_group_rank.cell(row=current_row_rank, column=col_num)
                cell.fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
            current_row_rank += 1
            
        last_group_rank = group_index
        
        username = acc.get('username', '')
        display_jindou = acc.get('display_jindou', max(acc.get('final_jindou', 0), acc.get('initial_jindou', 0)))
        actual_password = acc.get('actual_password', '')
        jindou_success = acc.get('jindou_success', False)
        password_error = acc.get('password_error', False)
        
        display_password = actual_password if actual_password else ''
        display_status = get_display_status(acc)
        
        if display_jindou > 0:
            year_end_prediction = calculate_year_end_prediction(int(display_jindou))
        else:
            year_end_prediction = ''
            
        global_rank = acc.get('global_rank', '')
        
        row_data = [
            global_rank, display_jindou, username, display_password,
            f"{group_index}组账号{acc.get('account_index', 0)}", display_status, year_end_prediction
        ]
        ws_group_rank.append(row_data)
        
        for col_num in range(1, total_cols + 1):
            cell = ws_group_rank.cell(row=current_row_rank, column=col_num)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if col_num == 1:
                cell.font = Font(bold=True)
            if col_num == 2:
                if display_jindou >= 500:
                    cell.font = Font(color="C00000", bold=True)
                elif display_jindou >= 300:
                    cell.font = Font(color="FF6600", bold=True)
                elif display_jindou >= 100:
                    cell.font = Font(color="0070C0", bold=True)
            if col_num == 6:
                if password_error:
                    cell.font = status_styles['password']
                elif jindou_success:
                    if display_status == '已签到过':
                        cell.font = status_styles['already']
                    else:
                        cell.font = status_styles['success']
                else:
                    cell.font = status_styles['fail']
            if col_num == 7 and isinstance(year_end_prediction, (int, float)):
                if year_end_prediction >= 1500:
                    cell.font = Font(color="C00000", bold=True)
                elif year_end_prediction >= 1000:
                    cell.font = Font(color="FF6600", bold=True)
                elif year_end_prediction >= 800:
                    cell.font = Font(color="0070C0", bold=True)
                    
        current_row_rank += 1

    # ==============================
    # 全局格式设置（列宽、边框、冻结窗格）
    # ==============================
    column_widths = [8, 12, 18, 15, 18, 18, 12]
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 格式化 Sheet 1
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=total_cols):
        for cell in row:
            cell.border = thin_border
    ws.freeze_panes = 'A2'
    
    # 格式化 Sheet 2
    for i, width in enumerate(column_widths, 1):
        ws_group.column_dimensions[get_column_letter(i)].width = width
    for row in ws_group.iter_rows(min_row=1, max_row=ws_group.max_row, min_col=1, max_col=total_cols):
        for cell in row:
            cell.border = thin_border
    ws_group.freeze_panes = 'A2'

    # 格式化 Sheet 3
    for i, width in enumerate(column_widths, 1):
        ws_group_rank.column_dimensions[get_column_letter(i)].width = width
    for row in ws_group_rank.iter_rows(min_row=1, max_row=ws_group_rank.max_row, min_col=1, max_col=total_cols):
        for cell in row:
            cell.border = thin_border
    ws_group_rank.freeze_panes = 'A2'
    
    # 保存文件
    wb.save(filename)
    log(f"Excel文件已生成: {filename}")
    
    return filename

def get_push_title():
    """获取推送标题"""
    now = datetime.now()
    return f"{now.month}月{now.day}日立创金豆签到结果"

def get_push_content():
    """获取推送内容"""
    now = datetime.now()
    month = now.month
    day = now.day
    
    failed_count = len(summary_info['failed_accounts'])
    pwd_error_count = len(summary_info['password_error_accounts'])
    
    if failed_count == 0 and pwd_error_count == 0:
        return f"{month}月{day}日立创金豆签到已全部成功"
    else:
        content_parts = []
        
        if failed_count > 0:
            # 按组分类失败账号
            failed_by_group = {}
            for acc in summary_info['failed_accounts']:
                group = acc['group']
                if group not in failed_by_group:
                    failed_by_group[group] = []
                failed_by_group[group].append(str(acc['index']))
            
            for group, indices in failed_by_group.items():
                content_parts.append(f"{group}组账号{','.join(indices)}")
        
        if pwd_error_count > 0:
            pwd_by_group = {}
            for acc in summary_info['password_error_accounts']:
                group = acc['group']
                if group not in pwd_by_group:
                    pwd_by_group[group] = []
                pwd_by_group[group].append(str(acc['index']))
            
            for group, indices in pwd_by_group.items():
                content_parts.append(f"{group}组账号{','.join(indices)}(密码错误)")
        
        return f"{month}月{day}日立创金豆签到有{'/'.join(content_parts)}失败"

def get_workflow_url():
    """获取GitHub Actions工作流运行页面链接"""
    server_url = os.getenv('GITHUB_SERVER_URL', 'https://github.com')
    repository = os.getenv('GITHUB_REPOSITORY', '')
    run_id = os.getenv('GITHUB_RUN_ID', '')
    
    if repository and run_id:
        return f"{server_url}/{repository}/actions/runs/{run_id}"
    return ""

def push_to_telegram(text, excel_file=None):
    """推送到Telegram"""
    bot_token = os.getenv('TELEGRAM_BOT_TOKEN')
    chat_id = os.getenv('TELEGRAM_CHAT_ID')
    
    if not bot_token or not chat_id:
        return False
    
    try:
        # 先发送文字消息
        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        params = {
            'chat_id': chat_id,
            'text': text,
            'parse_mode': 'HTML'
        }
        response = requests.post(url, params=params, timeout=30)
        
        if response.status_code == 200:
            log("Telegram-文字消息已推送")
        else:
            log(f"Telegram-文字推送失败: {response.text}")
            return False
        
        # 如果提供了Excel文件，单独发送文件
        if excel_file and os.path.exists(excel_file):
            doc_url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
            
            with open(excel_file, 'rb') as f:
                files = {'document': (os.path.basename(excel_file), f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                data = {'chat_id': chat_id}
                
                doc_response = requests.post(doc_url, data=data, files=files, timeout=60)
                
                if doc_response.status_code == 200:
                    log("Telegram-Excel文件已推送")
                    return True
                else:
                    log(f"Telegram-文件推送失败: {doc_response.text}")
                    return False
        
        return True
        
    except Exception as e:
        log(f"Telegram-推送异常: {e}")
        return False

def push_to_wechat(text, excel_file=None):
    """推送到企业微信"""
    webhook_key = os.getenv('WECHAT_WEBHOOK_KEY')
    
    if not webhook_key:
        return False
    
    try:
        # 先发送文字消息
        if webhook_key.startswith('https://'):
            url = webhook_key
        else:
            url = f"https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={webhook_key}"
        
        body = {"msgtype": "text", "text": {"content": text}}
        response = requests.post(url, json=body, timeout=30)
        
        if response.status_code == 200:
            resp_json = response.json()
            if resp_json.get('errcode') == 0:
                log("企业微信-文字消息已推送")
            else:
                log(f"企业微信-文字推送失败: {resp_json}")
                return False
        else:
            log(f"企业微信-文字推送失败: {response.text}")
            return False
        
        # 如果提供了Excel文件，先上传再发送
        if excel_file and os.path.exists(excel_file):
            # 从webhook URL中提取key
            if 'key=' in webhook_key:
                key = webhook_key.split('key=')[-1]
            else:
                key = webhook_key
            
            upload_url = f"https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key={key}&type=file"
            
            with open(excel_file, 'rb') as f:
                files = {'media': (os.path.basename(excel_file), f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                upload_response = requests.post(upload_url, files=files, timeout=60)
                
                if upload_response.status_code == 200:
                    upload_data = upload_response.json()
                    if upload_data.get('errcode') == 0:
                        media_id = upload_data.get('media_id')
                        
                        # 发送文件消息
                        send_url = f"https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={key}"
                        file_body = {
                            "msgtype": "file",
                            "file": {"media_id": media_id}
                        }
                        send_response = requests.post(send_url, json=file_body, timeout=30)
                        
                        if send_response.status_code == 200:
                            send_data = send_response.json()
                            if send_data.get('errcode') == 0:
                                log("企业微信-Excel文件已推送")
                                return True
                            else:
                                log(f"企业微信-文件发送失败: {send_data}")
                                return False
                        else:
                            log(f"企业微信-文件发送失败: {send_response.text}")
                            return False
                    else:
                        log(f"企业微信-文件上传失败: {upload_data}")
                        return False
                else:
                    log(f"企业微信-文件上传失败: {upload_response.text}")
                    return False
        
        return True
        
    except Exception as e:
        log(f"企业微信-推送异常: {e}")
        return False

def push_to_dingtalk(text):
    """推送到钉钉"""
    webhook = os.getenv('DINGTALK_WEBHOOK')
    
    if not webhook:
        return False
    
    try:
        if webhook.startswith('https://'):
            url = webhook
        else:
            url = f"https://oapi.dingtalk.com/robot/send?access_token={webhook}"
        
        body = {"msgtype": "text", "text": {"content": text}}
        response = requests.post(url, json=body, timeout=30)
        
        if response.status_code == 200:
            resp_json = response.json()
            if resp_json.get('errcode') == 0:
                log("钉钉-日志已推送")
                return True
            else:
                log(f"钉钉-推送失败: {resp_json}")
                return False
        else:
            log(f"钉钉-推送失败: {response.text}")
            return False
            
    except Exception as e:
        log(f"钉钉-推送异常: {e}")
        return False

def push_to_pushplus(text):
    """推送到PushPlus"""
    token = os.getenv('PUSHPLUS_TOKEN')
    
    if not token:
        return False
    
    try:
        url = "http://www.pushplus.plus/send"
        title = get_push_title()
        body = {"token": token, "title": title, "content": text}
        response = requests.post(url, json=body, timeout=30)
        
        if response.status_code == 200:
            log("PushPlus-日志已推送")
            return True
        else:
            log(f"PushPlus-推送失败: {response.text}")
            return False
            
    except Exception as e:
        log(f"PushPlus-推送异常: {e}")
        return False

def push_to_serverchan(text):
    """推送到Server酱"""
    sckey = os.getenv('SERVERCHAN_SCKEY')
    
    if not sckey:
        return False
    
    try:
        url = f"https://sctapi.ftqq.com/{sckey}.send"
        title = get_push_title()
        body = {"title": title, "desp": text}
        response = requests.post(url, data=body, timeout=30)
        
        if response.status_code == 200:
            log("Server酱-日志已推送")
            return True
        else:
            log(f"Server酱-推送失败: {response.text}")
            return False
            
    except Exception as e:
        log(f"Server酱-推送异常: {e}")
        return False

def push_to_serverchan3(text):
    """推送到Server酱3"""
    sckey = os.getenv('SERVERCHAN3_SCKEY')
    
    if not sckey:
        return False
    
    try:
        from serverchan_sdk import sc_send
        title = get_push_title()
        options = {"tags": "嘉立创|签到"}
        response = sc_send(sckey, title, text, options)
        
        if response.get("code") == 0:
            log("Server酱3-日志已推送")
            return True
        else:
            log(f"Server酱3-推送失败: {response}")
            return False
            
    except Exception as e:
        log(f"Server酱3-推送异常: {e}")
        return False

def push_to_coolpush(text):
    """推送到酷推"""
    skey = os.getenv('COOLPUSH_SKEY')
    
    if not skey:
        return False
    
    try:
        url = f"https://push.xuthus.cc/send/{skey}?c={text}"
        response = requests.get(url, timeout=30)
        
        if response.status_code == 200:
            log("酷推-日志已推送")
            return True
        else:
            log(f"酷推-推送失败: {response.text}")
            return False
            
    except Exception as e:
        log(f"酷推-推送异常: {e}")
        return False

def push_to_custom(text):
    """推送到自定义API"""
    webhook = os.getenv('CUSTOM_WEBHOOK')
    
    if not webhook:
        return False
    
    try:
        title = get_push_title()
        body = {"title": title, "content": text}
        response = requests.post(webhook, json=body, timeout=30)
        
        if response.status_code == 200:
            log("自定义API-日志已推送")
            return True
        else:
            log(f"自定义API-推送失败: {response.text}")
            return False
            
    except Exception as e:
        log(f"自定义API-推送异常: {e}")
        return False

def push_all_notifications(excel_file):
    """推送所有通知"""
    title = get_push_title()
    content = get_push_content()
    workflow_url = get_workflow_url()
    
    # 构建推送文本
    push_text = f"{title}\n\n{content}"
    
    if workflow_url:
        push_text += f"\n请访问以下链接，在Artifacts板块下载金豆详细排名：\n{workflow_url}"
    
    log(f"推送内容:\n{push_text}")
    
    # 检查是否有配置任何webhook
    has_webhook = any([
        os.getenv('TELEGRAM_BOT_TOKEN') and os.getenv('TELEGRAM_CHAT_ID'),
        os.getenv('WECHAT_WEBHOOK_KEY'),
        os.getenv('DINGTALK_WEBHOOK'),
        os.getenv('PUSHPLUS_TOKEN'),
        os.getenv('SERVERCHAN_SCKEY'),
        os.getenv('SERVERCHAN3_SCKEY'),
        os.getenv('COOLPUSH_SKEY'),
        os.getenv('CUSTOM_WEBHOOK')
    ])
    
    if not has_webhook:
        log("未配置任何webhook，跳过推送")
        return
    
    # 企业微信和Telegram需要单独处理文件上传
    # 其他平台只推送文字
    
    # 1. Telegram - 支持文件上传
    if os.getenv('TELEGRAM_BOT_TOKEN') and os.getenv('TELEGRAM_CHAT_ID'):
        push_to_telegram(push_text, excel_file)
    
    # 2. 企业微信 - 支持文件上传
    if os.getenv('WECHAT_WEBHOOK_KEY'):
        push_to_wechat(push_text, excel_file)
    
    # 3. 钉钉 - 仅文字
    push_to_dingtalk(push_text)
    
    # 4. PushPlus - 仅文字
    push_to_pushplus(push_text)
    
    # 5. Server酱 - 仅文字
    push_to_serverchan(push_text)
    
    # 6. Server酱3 - 仅文字
    push_to_serverchan3(push_text)
    
    # 7. 酷推 - 仅文字
    push_to_coolpush(push_text)
    
    # 8. 自定义API - 仅文字
    push_to_custom(push_text)

def clean_temp_files():
    """清理临时结果文件"""
    temp_files = glob.glob('jlc_result_*.json')
    
    for filepath in temp_files:
        try:
            os.remove(filepath)
            log(f"已删除临时文件: {filepath}")
        except Exception as e:
            log(f"删除文件 {filepath} 失败: {e}")

def main():
    log("开始汇总金豆签到结果...")
    
    # 1. 收集所有结果
    if not collect_all_results():
        log("收集结果失败，退出")
        sys.exit(1)
    
    if not all_accounts:
        log("没有账号数据，退出")
        sys.exit(1)
    
    log(f"共收集到 {len(all_accounts)} 个账号的数据")
    
    # 2. 生成Excel文件
    excel_file = generate_excel()
    
    # 3. 推送通知
    push_all_notifications(excel_file)
    
    # 4. 清理临时文件
    clean_temp_files()
    
    log("汇总完成!")

if __name__ == "__main__":
    main()
